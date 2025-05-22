import os
import json
from chord_extractor.extractors import Chordino
from difflib import SequenceMatcher
import json
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.cell.rich_text import CellRichText, TextBlock, InlineFont
import autochord
import csv
from pathlib import Path

# Set up parameters
xlsx_file = "JM.xlsx"
audio_folder = "Joyful_Melodies"
transcripts_folder = "transcripts"  # or whatever path you're using for the transcripts
output_folder = "aligned_output"
chord_output_folder = "chord_transcriptions"

# Ensure output folders exist
os.makedirs(output_folder, exist_ok=True)
os.makedirs(chord_output_folder, exist_ok=True)

def read_lyrics_from_xlsx(file_path, serial_number):
    """Read lyrics from the XLSX file for the given serial number."""
    wb = load_workbook(file_path, rich_text=True)
    sheet = wb.active
    
    for row in sheet.iter_rows(min_row=1, values_only=False):
        if row[0].value == serial_number:
            title = row[1].value
            lyrics_cell = row[2]
            
            chorus = []
            verses = []
            current_verse = []
            
            if isinstance(lyrics_cell.value, CellRichText):
                lines = str(lyrics_cell.value).split('\n')
                for line in lines:
                    if line.strip():
                        is_bold = any(isinstance(run, TextBlock) and run.font and run.font.b 
                                      for run in lyrics_cell.value if str(run) in line)
                        if is_bold:
                            chorus.append(line.strip())
                        else:
                            current_verse.append(line.strip())
                    elif current_verse:
                        verses.append('\n'.join(current_verse))
                        current_verse = []
            else:
                # Fallback to non-rich text handling
                lines = str(lyrics_cell.value).split('\n')
                for line in lines:
                    if line.strip():
                        current_verse.append(line.strip())
                    elif current_verse:
                        verses.append('\n'.join(current_verse))
                        current_verse = []
            
            if current_verse:
                verses.append('\n'.join(current_verse))
            
            return {
                "title": title,
                "chorus": '\n'.join(chorus),
                "verses": verses
            }
    
    return None

def extract_and_combine_chords(audio_file):
    autochord_chords = autochord.recognize(audio_file)
    chordino = Chordino()
    chordino_chords = chordino.extract(audio_file)
    
    combined_chords = []
    chordino_index = 0
    
    for i, (start, end, autochord_chord) in enumerate(autochord_chords):
        if autochord_chord == 'N':
            while chordino_index < len(chordino_chords) and chordino_chords[chordino_index].timestamp < end:
                if chordino_chords[chordino_index].timestamp >= start:
                    converted_chord = chordino_chords[chordino_index].chord
                    break
                chordino_index += 1
            else:
                continue
        elif ':' in autochord_chord:
            chord_root, chord_type = autochord_chord.split(':')
            converted_chord = chord_root if chord_type == 'maj' else f"{chord_root}m"
        else:
            converted_chord = autochord_chord
        
        while chordino_index < len(chordino_chords) and chordino_chords[chordino_index].timestamp < end:
            chordino_chord = chordino_chords[chordino_index].chord
            if chordino_chord[0] == converted_chord[0]:
                converted_chord = chordino_chord
                break
            chordino_index += 1
        
        # Adjust end time to the start of the next chord
        if i < len(autochord_chords) - 1:
            end = autochord_chords[i+1][0]
        
        combined_chords.append((start, end, converted_chord))
    
    return combined_chords

def read_transcription(transcription_file):
    """Read the transcription file and return a list of (word, start_time, end_time) tuples."""
    transcription = []
    with open(transcription_file, 'r') as f:
        for line in f:
            parts = line.strip().split()
            if len(parts) >= 4:
                word = parts[0]
                start_time = float(parts[1][1:-1])  # Remove parentheses
                end_time = float(parts[3][:-1])  # Remove parenthesis
                transcription.append((word, start_time, end_time))
    return transcription


def correct_transcription(lyrics, transcription):
    """Corrects the transcription using the lyrics from the XLSX file."""
    lyrics_words = ' '.join(lyrics["chorus"].lower().split() + 
                            [word for verse in lyrics["verses"] for word in verse.lower().split()])
    lyrics_words = lyrics_words.split()
    
    corrected_transcription = []
    lyrics_index = 0
    window_size = 5  # Look at 5 words before and after the current word
    
    for i, (trans_word, start, end) in enumerate(transcription):
        trans_word = trans_word.lower()
        
        # Define the search window in lyrics
        window_start = max(0, lyrics_index - window_size)
        window_end = min(len(lyrics_words), lyrics_index + window_size + 1)
        search_window = lyrics_words[window_start:window_end]
        
        # Find the best matching word from lyrics within the window
        best_match = None
        best_ratio = 0
        for j, lyric_word in enumerate(search_window):
            ratio = SequenceMatcher(None, trans_word, lyric_word).ratio()
            if ratio > best_ratio and ratio > 0.8:  # Increased similarity threshold
                best_ratio = ratio
                best_match = lyric_word
                lyrics_index = window_start + j
        
        if best_match:
            corrected_transcription.append((best_match, start, end))
            lyrics_index += 1
        else:
            # If no good match found, keep the original word
            corrected_transcription.append((trans_word, start, end))
    
    return corrected_transcription

def align_lyrics_and_chords(corrected_transcription, chords, lyrics):
    """Aligns lyrics with corrected transcription and chords, preserving original line structure."""
    aligned_output = []
    chord_index = 0
    
    # Flatten lyrics into lines
    lyric_lines = [line.strip() for verse in lyrics['verses'] for line in verse.split('\n') if line.strip()]
    if lyrics['chorus']:
        lyric_lines = lyrics['chorus'].split('\n') + lyric_lines
    
    word_count = 0
    for lyric_line in lyric_lines:
        line_words = lyric_line.split()
        current_line = []
        last_chord = None
        
        for i, word in enumerate(line_words):
            if word_count >= len(corrected_transcription):
                break
            
            _, start, end = corrected_transcription[word_count]
            
            # Find the chord that occurs just before or at the start of the word
            while chord_index < len(chords) - 1 and chords[chord_index + 1][0] <= start:
                chord_index += 1
            
            current_chord = chords[chord_index][2]  # Use the actual chord name
            chord_start = chords[chord_index][0]
            next_chord_start = chords[chord_index + 1][0] if chord_index < len(chords) - 1 else float('inf')
            
            # Decide whether to place the chord before or after the word
            if abs(chord_start - start) < abs(chord_start - end) and chord_start > start:
                word_with_chord = f"{word}[{current_chord}]"
            elif current_chord != last_chord:
                word_with_chord = f"[{current_chord}]{word}"
                last_chord = current_chord
            else:
                word_with_chord = word
            
            current_line.append(word_with_chord)
            word_count += 1
        
        aligned_output.append(" ".join(current_line))
    
    return aligned_output

def save_chord_transcription(chords, output_file):
    """Save the chord transcription to a file."""
    with open(output_file, 'w') as f:
        for start, end, chord in chords:
            f.write(f"{chord} ({start:.2f} - {end:.2f})\n")

def calculate_error_rate(transcription, corrected_transcription):
    total_words = len(transcription)
    errors = sum(1 for orig, corr in zip(transcription, corrected_transcription) if orig[0].lower() != corr[0].lower())
    error_rate = errors / total_words
    return error_rate

def process_song(serial_number, prefix):
    print(f"\nProcessing song {serial_number}")
    
    analysis_data = {
        "Serial Number": serial_number,
        "Prefix": prefix,
        "Title": "",
        "Transcription Length": 0,
        "Transcription Error Rate": 0.0
    }
    
    try:
        # Read lyrics
        lyrics = read_lyrics_from_xlsx(xlsx_file, serial_number)
        if lyrics is None:
            print(f"No lyrics found for serial number {serial_number}")
            return
        print(f"Lyrics read successfully. Title: {lyrics['title']}")
        analysis_data["Title"] = lyrics['title']
        
        # Read pre-extracted transcription
        transcription_file = os.path.join(transcripts_folder, f"{prefix}_{serial_number:03d}_transcription.txt")
        if not os.path.exists(transcription_file):
            print(f"No transcription file found for serial number {serial_number}")
            return
        transcription = read_transcription(transcription_file)
        print(f"Transcription length: {len(transcription)}")
        analysis_data["Transcription Length"] = len(transcription)
        
        # Correct the transcription
        corrected_transcription = correct_transcription(lyrics, transcription)
        
        # Calculate error rate
        error_rate = calculate_error_rate(transcription, corrected_transcription)
        print(f"Transcription error rate: {error_rate:.2%}")
        analysis_data["Transcription Error Rate"] = error_rate
        
        # Find audio file and extract chords
        audio_files = [f for f in os.listdir(audio_folder) if f.startswith(f"{serial_number:03d}")]
        if not audio_files:
            print(f"No audio file found for serial number {serial_number}")
            return
        audio_file = os.path.join(audio_folder, audio_files[0])
        chords = extract_and_combine_chords(audio_file)
        print(f"Extracted {len(chords)} chords")
        
        # Save combined chord transcription
        chord_output_file = os.path.join(chord_output_folder, f"{prefix}_{serial_number:03d}_chords.txt")
        save_chord_transcription(chords, chord_output_file)
        print(f"Combined chord transcription saved to {chord_output_file}")
        
        # Align lyrics, corrected transcription, and chords
        aligned_output = align_lyrics_and_chords(corrected_transcription, chords, lyrics)
        
        # Save aligned output
        aligned_output_file = os.path.join(output_folder, f"{prefix}_{serial_number:03d}_aligned.txt")
        with open(aligned_output_file, 'w') as f:
            f.write('\n'.join(aligned_output))
        
        print(f"Aligned output saved to {aligned_output_file}")
        
        # Append analysis data to CSV
        append_to_analysis_csv(analysis_data)
        
    except Exception as e:
        print(f"Error processing song {serial_number}: {str(e)}")
        import traceback
        traceback.print_exc()

def append_to_analysis_csv(data):
    csv_file = Path("analysis.csv")
    file_exists = csv_file.is_file()
    
    with open(csv_file, 'a', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=["Serial Number", "Prefix", "Title", "Transcription Length", "Transcription Error Rate"])
        
        if not file_exists:
            writer.writeheader()
        
        writer.writerow(data)

# Process Western and Indian songs
#western_serials = [4, 5, 27, 60, 64, 74, 90, 134, 149, 194, 312, 325,375, 424, 581,601, 615, 691]
#indian_serials = [1, 3, 14, 67, 77, 87, 133, 161, 188, 266, 353, 362, 406, 547, 657, 678]
western_serials = [503, 511, 514, 537, 613, 685, 375, 361, 283, 651 ]
indian_serials = [ 509, 517,  544, 564, 569, 434, 423, 338, 381]

for serial in western_serials:
    process_song(serial, "Western")

for serial in indian_serials:
    process_song(serial, "Indian")

print("All songs processed.")
